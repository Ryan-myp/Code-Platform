"""短剧工厂 v15 单元测试：分镜表导出 Excel + 批量素材清单。

纯函数级测试（build_shot_sheet / build_material_manifest）+ 端点集成测试。
"""

import io
import json
import sys
from pathlib import Path

BACKEND = str(Path(__file__).resolve().parents[2] / "backend")
if BACKEND not in sys.path:
    sys.path.insert(0, BACKEND)

from short_drama import build_material_manifest, build_shot_sheet  # noqa: E402


def _scenes() -> list[dict]:
    return [
        {
            "id": 1,
            "chars": ["lin"],
            "shot": "雨夜小巷，主角撑伞快步走过",
            "search": "night city rain",
            "narrator": "深夜的城市，故事开始了",
            "dialogue": "我终于找到了这里",
            "emotion": "serious",
            "sec": 8,
        },
        {
            "id": 2,
            "chars": ["lin", "wang"],
            "shot": "老旧咖啡馆内景",
            "search": "cozy cafe interior, coffee",
            "narrator": "这里，是一切开始的地方",
            "dialogue": "好久不见",
            "emotion": "gentle",
            "sec": 6,
        },
        {"id": 3, "chars": [], "shot": "城市全景航拍", "search": "", "narrator": "", "dialogue": "", "emotion": "neutral", "sec": 5},
    ]


def _characters() -> list[dict]:
    return [
        {"id": "lin", "name": "林小满", "gender": "女", "age": "24岁"},
        {"id": "wang", "name": "王警官", "gender": "男", "age": "35岁"},
    ]


class TestBuildShotSheet:
    """分镜表 xlsx：结构 / 表头 / 角色名映射 / 情绪中文。"""

    def _load(self, data: bytes):
        import openpyxl

        return openpyxl.load_workbook(io.BytesIO(data))

    def test_sheet_structure(self):
        data = build_shot_sheet(_scenes(), "雨夜追凶", _characters())
        wb = self._load(data)
        assert "分镜表" in wb.sheetnames
        ws = wb["分镜表"]
        # 标题行 + 空行 + 表头 + 3 镜 = 6 行
        assert ws.max_row == 6
        assert ws.cell(3, 1).value == "镜号"
        assert ws.cell(3, 5).value == "画面描述(shot)"

    def test_title_row_merged(self):
        data = build_shot_sheet(_scenes(), "雨夜追凶")
        wb = self._load(data)
        ws = wb["分镜表"]
        assert ws.cell(1, 1).value == "《雨夜追凶》分镜表"
        assert ws.merged_cells.ranges  # 标题跨列合并

    def test_character_names_mapped(self):
        data = build_shot_sheet(_scenes(), "", _characters())
        wb = self._load(data)
        ws = wb["分镜表"]
        row2 = [ws.cell(2, c).value for c in range(1, 9)]
        assert row2[3] == "林小满"  # 出场角色列
        assert [ws.cell(3, c).value for c in range(1, 9)][3] == "林小满、王警官"

    def test_emotion_chinese(self):
        data = build_shot_sheet(_scenes())
        wb = self._load(data)
        ws = wb["分镜表"]
        assert ws.cell(2, 3).value == "严肃"
        assert ws.cell(3, 3).value == "温柔"

    def test_empty_scenes_still_headers(self):
        data = build_shot_sheet([])
        wb = self._load(data)
        ws = wb["分镜表"]
        assert ws.cell(1, 1).value == "镜号"
        assert ws.max_row == 1

    def test_sec_and_text_filled(self):
        data = build_shot_sheet(_scenes())
        wb = self._load(data)
        ws = wb["分镜表"]
        assert ws.cell(2, 2).value == 8
        assert "深夜的城市" in ws.cell(2, 7).value


class TestBuildMaterialManifest:
    """素材清单：关键词回退 / 汇总统计 / md 段落。"""

    def test_items_structure(self):
        result = build_material_manifest(_scenes())
        assert len(result["items"]) == 3
        it = result["items"][0]
        assert it["no"] == 1 and it["keyword"] == "night city rain"
        assert it["sec"] == 8 and it["emotion"] == "严肃"
        assert it["text_len"] > 0

    def test_keyword_fallback_to_shot(self):
        result = build_material_manifest(_scenes())
        assert result["items"][2]["keyword"] == "城市全景航拍"

    def test_keyword_fallback_scene_no(self):
        result = build_material_manifest([{"shot": "", "search": "", "narrator": "", "dialogue": "", "sec": 5}])
        assert result["items"][0]["keyword"] == "scene_1"

    def test_summary_stats(self):
        result = build_material_manifest(_scenes())
        assert result["summary"]["total_scenes"] == 3
        assert result["summary"]["total_sec"] == 19
        assert result["summary"]["total_text_chars"] > 0
        # 关键词拆分去重（night/city/rain/cozy/cafe/interior/coffee + shot 回退）
        assert "night" in result["summary"]["keywords"]
        assert "coffee" in result["summary"]["keywords"]

    def test_manifest_md_sections(self):
        result = build_material_manifest(_scenes())
        md = result["manifest_md"]
        assert "# 短剧素材清单" in md
        assert "| 镜号 | 素材关键词 |" in md
        assert "## 关键词汇总" in md and "## 使用说明" in md
        assert "- night" in md

    def test_empty_scenes(self):
        result = build_material_manifest([])
        assert result["items"] == []
        assert result["summary"]["total_scenes"] == 0
        assert "- （无）" in result["manifest_md"]


class TestManifestEndpoint:
    """material-manifest / export-shots 端点：200 正常 / 400 非法输入。"""

    def _client(self, auth_headers):
        from fastapi.testclient import TestClient

        from main import app

        return TestClient(app)

    def test_manifest_ok(self, auth_headers):
        client = self._client(auth_headers)
        resp = client.post(
            "/api/drama/material-manifest",
            data={"scenes_json": json.dumps(_scenes(), ensure_ascii=False)},
            headers=auth_headers,
        )
        assert resp.status_code == 200
        data = resp.json()
        assert data["summary"]["total_scenes"] == 3
        assert "manifest_md" in data

    def test_manifest_bad_json(self, auth_headers):
        client = self._client(auth_headers)
        resp = client.post("/api/drama/material-manifest", data={"scenes_json": "{bad"}, headers=auth_headers)
        assert resp.status_code == 400
        assert "分镜 JSON" in resp.json()["detail"]

    def test_manifest_empty_scenes(self, auth_headers):
        client = self._client(auth_headers)
        resp = client.post("/api/drama/material-manifest", data={"scenes_json": "[]"}, headers=auth_headers)
        assert resp.status_code == 400

    def test_export_shots_ok(self, auth_headers):
        client = self._client(auth_headers)
        resp = client.post(
            "/api/drama/export-shots",
            data={
                "title": "雨夜追凶",
                "scenes_json": json.dumps(_scenes(), ensure_ascii=False),
                "characters_json": json.dumps(_characters(), ensure_ascii=False),
            },
            headers=auth_headers,
        )
        assert resp.status_code == 200
        assert resp.headers["content-type"].startswith("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        assert len(resp.content) > 1000
        # 产物可被 openpyxl 回读
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        assert "分镜表" in wb.sheetnames

    def test_export_shots_bad_json(self, auth_headers):
        client = self._client(auth_headers)
        resp = client.post("/api/drama/export-shots", data={"scenes_json": "[1,2"}, headers=auth_headers)
        assert resp.status_code == 400
